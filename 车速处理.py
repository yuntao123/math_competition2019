
# (2)汽车加、减速度异常的数据（普通轿车一般情况下：0至100km/h的加速时间大于7秒，紧急刹车最大减速度在7.5~8 m/s2）；

# 处理思路：先处理最大减速度，在处理1-100KM/s
import numpy as np
import matplotlib.pyplot as plt
from data import get_data
import xlwt
import openpyxl

outwb = openpyxl.Workbook()  # 打开一个将写的文件
outws = outwb.create_sheet(index=0)  # 在将写的文件创建she

x= get_data(1,2)  # 提取车速
print(x[1][0])


def speed_process(): # 设置初速度
    v0 = x[0][0]
    erro_num = 0
    a = []
    for i in range(1,len(x)):
        v = x[i][0]
        a.append(v-v0)
        if v-v0 < -8:
            erro_num += 1
            v = v0-8
            outws.cell(i + 1, 1).value = v  # 写文件
        else:
            outws.cell(i+2, 1).value = v  # 写文件
        v0 = v
    print('最大减速度：',min(a))
    print('error_num',erro_num)
    ####################################################################################################################
    # 保存该excel文件,有同名文件时直接覆盖.............................................
    outws.cell(1, 1).value = ('GPS车速')  # 写文件
    outws.cell(2, 1).value = (v0)  # 写文件

def aceer_pre():
    speed_win = []
    speed_eval = []
    err_spe,i,j =0,0,0
    for i in range(len(x)):
        speed_win.clear()
        if i + j <= len(x) - 1:
            for j in range(7):
                speed_win.append(x[i+j][0])
                speed_eval.append(max(speed_win) - min(speed_win))
            if max(speed_win) - min(speed_win) > 100:
                err_spe += 1
                speed_win[np.argmax(speed_win)] = min(speed_win) +100
            for k in range(7):
                outws.cell(i + 2+k, 2).value = speed_win[k]  # 写文件
    print('7S内最大速度差值：',max(speed_eval))
    print('0至100km/h的加速时间大于7秒个数：',err_spe)
    outws.cell(1, 2).value = ('GPS车速')  # 写文件

if __name__ == '__main__':
    aceer_pre()
    speed_process()
    outwb.save("./data_set/spped.xlsx")
    print('创建excel文件完成！')