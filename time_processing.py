import numpy as np
import matplotlib.pyplot as plt
from data import get_data
import xlwt
import openpyxl
# workbook = xlwt.Workbook() # 注意Workbook的开头W要大写
# sheet1 = workbook.add_sheet('sheet1',cell_overwrite_ok=True)

outwb = openpyxl.Workbook()  # 打开一个将写的文件
outws = outwb.create_sheet(index=0)  # 在将写的文件创建she

x= get_data(1)
s = (x[1][0]).split(':')  # 返回一个列表 s【2】
print(len(x))

def time_remedy(second,minu,hou,day):
    error_num = 0
    erro_list = []
    for i in range(1,len(x)):
        outws.cell(i+1, 1).value = i  # 写文件
        s = (x[i][0]).split(':')  # 返回一个列表 s【2】
        print('ssssss',s)
        sec_t = int(((s[-1]).split('.'))[0])
        min_t = int(s[-2])
        hou_t = int(((s[-3]).split(' '))[-1])
        day_t = int(((((s[0]).split('/'))[-1]).split(' '))[-2])
        print('day_t,hou_t',day_t,hou_t)
        if sec_t ==0:  # 秒测试1###################################################################################
            if second + 1 !=60:
                error_num += 1
                erro_list.append(i)
        elif second + 1 !=sec_t:   # 秒测试2############################################################################
            error_num += 1
            erro_list.append(i)
        second = sec_t # 重新赋值，进行下一轮

        if sec_t == 0:
            if min_t ==0:  # 分测试1###################################################################################
                if minu + 1 !=60:
                    error_num += 1
                    erro_list.append(i)
            elif minu + 1 !=min_t:   # 分测试2############################################################################
                error_num += 1
                erro_list.append(i)
        minu = min_t # 重新赋值，进行下一轮

        if min_t == 0:
            if hou_t == 0:  # 分测试1###################################################################################
                if hou + 1 != 24:
                    error_num += 1
                    erro_list.append(i)
            elif hou + 1 != hou_t:  # 分测试2############################################################################
                error_num += 1
                erro_list.append(i)
        hou = hou_t  # 重新赋值，进行下一轮

        if hou_t == 0:
            if day + 1 != day_t:  # 分测试2############################################################################
                error_num += 1
                erro_list.append(i)
        day = day_t  # 重新赋值，进行下一轮

        time_str = str("2017/12") + str(sec_t)
        outws.cell(i+1, 2).value = (time_str)  # 写文件
        print(s[-1])

    print('error_num:',error_num) #  打印错误数目

    for j in range(1,len(erro_list)):
        outws.cell(j, 3).value = erro_list[j-1]  # 写文件
    # 保存该excel文件,有同名文件时直接覆盖.............................................
    outws.cell(1, 1).value = (0)  # 写文件
    outws.cell(1, 2).value = (13)  # 写文件
    outwb.save("./data_set/time.xlsx")
    print('创建excel文件完成！')
time_remedy(second=13,minu =42,hou=13,day=18)
