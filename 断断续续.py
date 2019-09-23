# (4)长时间堵车、断断续续低速行驶情况（最高车速小于10km/h）,通常可按怠速情况处理。
# 怠速：汽车停止运动，但发动机保持最低转速运转的连续过程。
# 处理：最高车速小于10km/h发动机最低速，速度为0，油门踏板开度0
import numpy as np
import matplotlib.pyplot as plt
from data import get_data
import xlwt
import openpyxl

outwb = openpyxl.Workbook()  # 打开一个将写的文件
outws = outwb.create_sheet(index=0)  # 在将写的文件创建she

x1= get_data(0,1)  # 时间
x2= get_data(1,2)  # GPS车速
x3= get_data(2,3)  # X轴加速度
x4= get_data(3,4)  # Y轴加速度
x5= get_data(4,5)  # Z轴加速度
x6= get_data(5,6)  # 经度
x7= get_data(6,7)  # 纬度
x8= get_data(7,8)  # 发动机转速
x9= get_data(8,9)  # 扭矩百分比
x10= get_data(9,10)  # 瞬时油耗
x11= get_data(10,11)  # 油门踏板开度
x12= get_data(11,12)  # 空燃比
x13= get_data(12,13)  # 发动机负荷百分比
x14= get_data(13,14)  # 进气流量



def interval_speed(speed = 10,x = x2):
    count_0 =0
    count_list = []
    time_save = []
    staus = 0 # 写文件是否要的标志
    for i in range(len(x)):
        if x[i][0] <= speed:
            outws.cell(i + 2, 1).value = 0  # 写文件
            outws.cell(i + 2, 2).value = 162  # 写文件 发动机转速
            outws.cell(i + 2, 3).value = 0  # 脚踏开关
        else:
            outws.cell(i + 2, 1).value = x[i][0]  # 写文件
            outws.cell(i + 2, 2).value = x8[i][0]  # 写文件 发动机转速
            outws.cell(i + 2, 3).value = x11[i][0]  # 脚踏开关

        count_0 += 1
        print(count_0)

    outws.cell( 1, 1).value = "车速"  # 写文件
    outws.cell(1, 2).value = "发动机转速"  # 写文件
    outws.cell(1, 3).value = "脚踏开关"  # 写文件
    outwb.save("./data_set/断断续续.xlsx")
    print('创建excel文件完成！')

interval_speed(speed = 10,x = x2)