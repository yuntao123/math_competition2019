# 运动学片段是指汽车从怠速状态开始至下一个怠速状态开始之间的车速区间，如图3所示（基于运动学片段构建汽车行驶工况曲线是日前最常用的方法之一，
# 但并不是必须的步骤，有些构建汽车行驶工况曲线的方法并不需要进行运动学片段划分和提取）。请设计合理的方法，将上述经处理后的数据划分为多个运动学片段，
# 并给出各数据文件最终得到的运动学片段数量。
# 解决措施：通过观察停车时间确定一个合理的值


from data import get_data
import openpyxl

outwb = openpyxl.Workbook()  # 打开一个将写的文件
outws = outwb.create_sheet(index=0)  # 在将写的文件创建she

x1= get_data(0,1)  # 提取时间
x2= get_data(1,2)  # 提取车速
x3= get_data(2,3)  # 提取车速
x4= get_data(3,4)  # 提取车速
x5= get_data(4,5)  # 提取车速
x6= get_data(5,6)  # 提取车速
x7= get_data(6,7)  # 提取车速
x8= get_data(7,8)  # 提取车速
x9= get_data(8,9)  # 提取车速
x10= get_data(9,10)  # 提取车速
x11= get_data(10,11)  # 提取车速
x12= get_data(11,12)  # 提取车速
x13= get_data(12,13)  # 提取车速
x14= get_data(13,14)  # 提取车速

def window(x = x2,t_0=10):
    count_0 = 0
    count_1 = 0
    count_T = 0
    pianduan = []
    status = 0
    row =0
    for i in range(len(x)):
        pianduan.append(x[i][0])
        if x[i][0] == 0:
            count_0 += 1
            count_1 =0
        if x[i][0] != 0:
            count_1 += 1
            print('count_1',count_1)
            if i+1<len(x):
                if x[i+1][0] == 0:
                    if count_0 > t_0 and count_1> t_0:
                        row += 1
                        count_1 = 0
                        count_0 = 0
                        outws.cell(1, row).value = '第'+ str(row) + '片段' # 写文件
                        print('运动学片段',row)
                        for k in range(len(pianduan)):
                            outws.cell(k+2, row).value = pianduan[k]  # 写文件
                        pianduan.clear()
                    else:
                        pianduan.clear()
                        count_0 = 0

    #
    # outws.cell(1, 1).value = ('运动片段')  # 写文件
    # print('采集的运动学片段个数：',count_T)
if __name__ == '__main__':
    window(x=x2, t_0=10)
    outwb.save("./data_set/pianduan.xlsx")
    print('创建excel文件完成！')

