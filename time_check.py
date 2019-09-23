import numpy as np
import matplotlib.pyplot as plt
from data import get_data
import xlwt
import openpyxl
# workbook = xlwt.Workbook() # 注意Workbook的开头W要大写
# sheet1 = workbook.add_sheet('sheet1',cell_overwrite_ok=True)

outwb = openpyxl.Workbook()  # 打开一个将写的文件
outws = outwb.create_sheet(index=0)  # 在将写的文件创建she

x= get_data(0,1)

print(len(x))

def write_data(second,minu,hour,day):
    error_num = 0
    erro_list = []
    for i in range(len(x)):
        second += 1
        if second==60:
            second = 0
            minu += 1
            if minu == 60:
                minu =0
                hour += 1
                if hour == 24 :
                    hour = 0
                    day +=1
        data = str("2017/12/")+str(day) + str(' ') + str(hour)+str(':')+str(minu)+str(':')+str(second)
        outws.cell(i+2, 1).value = (data)  # 写文件

        s = (x[i][0]).split(':')  # 返回一个列表 s【2】
        # print('ssssss', s)
        sec_t = int(((s[-1]).split('.'))[0])
        min_t = int(s[-2])
        hou_t = int(((s[-3]).split(' '))[-1])
        day_t = int(((((s[0]).split('/'))[-1]).split(' '))[-2])
        if second !=sec_t:
            error_num +=1
            erro_list.append(i)
            print('second',x[i][0],data,i)
        else:
            if min_t != minu:
                error_num += 1
                erro_list.append(i)
                print('second', x[i][0], data,i)
            else:
                if hou_t != hour:
                    error_num += 1
                    erro_list.append(i)
                    print('second', x[i][0], data,i)
                else:
                    if day_t != day:
                        error_num += 1
                        erro_list.append(i)
                        print('second', x[i][0], data,i)

    print(erro_list)
    print('error_num:',error_num)

    for j in range(len(erro_list)):
        outws.cell(1, 2).value = ('error_index')  # 写文件
        outws.cell(j+2, 2).value = (erro_list[j])  # 写文件
    # 保存该excel文件,有同名文件时直接覆盖.............................................
    outws.cell(1, 1).value = ('时间')  # 写文件
    # outws.cell(2, 1).value = ("2017/12/18 13:42:13")  # 写文件
    outwb.save("./data_set/time.xlsx")
    print('创建excel文件完成！')
write_data(second=12,minu =42,hour=13,day=18)
