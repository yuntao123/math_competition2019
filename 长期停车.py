# (3)长期停车（如停车不熄火等候人、停车熄火了但采集设备仍在运行等）所采集的异常数据。
# 解决措施：通过观察停车时间确定一个合理的值

import numpy as np
import matplotlib.pyplot as plt
from data import get_data
import xlwt
import openpyxl

outwb = openpyxl.Workbook()  # 打开一个将写的文件
outws = outwb.create_sheet(index=0)  # 在将写的文件创建she
x1= get_data(0,1)  # 提取车速
x2= get_data(1,2)  # 提取车速
x3= get_data(2,3)  # 提取车速
x4= get_data(3,4)  # 提取车速
x5= get_data(4,5)  # 提取车速
x6= get_data(5,6)  # 提取车速
x7= get_data(6,7)  # 提取车速
x8= get_data(7,8)  # 提取车速
x9= get_data(8,9)  # 提取车速
x10= get_data(9,10)  # 提取车速
x11= get_data(10,11)  # 提取车速
x12= get_data(11,12)  # 提取车速
x13= get_data(12,13)  # 提取车速
x14= get_data(13,14)  # 提取车速

def zero_count(time = 30):
    count_0 =0
    count_list = []
    for i in range(len(x2)):
        if i +1 < len(x2):
            if x2[i][0] == 0:
                count_0 += 1
                if x2[i+1][0] != 0:
                    if count_0 >time:
                        count_list.append(count_0)
            else:
                count_0 = 0
    for j in range(len(count_list)):
        outws.cell(j+1, 1).value = count_list[j]  # 写文件
    outws.cell( 1, 1).value = "停车时间"  # 写文件
    outwb.save("./data_set/stop_time.xlsx")
    print('创建excel文件完成！')


def zero_delet(time = 180,x = x2):
    count_0 =0
    count_list = []
    time_save = []
    staus = 0 # 写文件是否要的标志
    for i in range(len(x)):
        time_save.clear()  #############清除内容#########################################################################
        if i +time < len(x):
            for j in range(time):
                time_save.append(x[i + j][0])
                if x[i+j][0] != 0: # 不满足停车时间
                    staus = 1
                else:
                    staus = 0
            if staus == 1:
                for k in range(len(time_save)):
                    # outws.cell(i+k + 1, 1).value = x1[i+k]  # 写文件
                    outws.cell(i + k + 1, 2).value = time_save[k]  # 写文件
                    outws.cell(i + k + 1, 3).value = x3[i+k][0]  # 写文件
                    outws.cell(i + k + 1, 4).value = x4[i+k][0]  # 写文件
                    outws.cell(i + k + 1, 5).value = x5[i+k][0]  # 写文件
                    outws.cell(i + k + 1, 6).value = x6[i+k][0]  # 写文件
                    outws.cell(i + k + 1, 7).value = x7[i+k][0]  # 写文件
                    outws.cell(i + k + 1, 8).value = x8[i+k][0]  # 写文件
                    outws.cell(i + k + 1, 9).value = x9[i+k][0]  # 写文件
                    outws.cell(i + k + 1, 10).value = x10[i+k][0]  # 写文件
                    outws.cell(i + k + 1, 11).value = x11[i+k][0]  # 写文件
                    outws.cell(i + k + 1, 12).value = x12[i+k][0]  # 写文件
                    outws.cell(i + k + 1, 13).value = x13[i+k][0]  # 写文件
                    outws.cell(i + k + 1, 14).value = x14[i+k][0]  # 写文件
                    count_0 += 1
                    print(count_0)

    outws.cell( 1, 1).value = "停车时间"  # 写文件
    outwb.save("./data_set/zero_delet.xlsx")
    print('创建excel文件完成！')

zero_delet(time = 180,x = x2)

# zero_count()