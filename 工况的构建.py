# 请根据上述经处理后的数据，构建一条能体现参与数据采集汽车行驶特征的汽车行驶工况曲线（1200-1300秒），该曲线的汽车运动特征能代表所采集数
# 据源（经处理后的数据）的相应特征，两者间的误差越小，说明所构建的汽车行驶工况的代表性越好。要求：
# （1）科学、有效的构建方法（数学模型或算法，特别鼓励创新方法，如果采用已有的方法，必须注明来源）；
# （2）合理的汽车运动特征评估体系（至少包含但不限于以下指标：平均速度（km/h）、平均行驶速度（km/h）、平均加速度（m/）、平均减速度（m/）、
# 怠速时间比（%）、加速时间比（%）、减速时间比（%）、速度标准差（km/h）、加速度标准差（m/）等）；
# （3）按照你们所构建的汽车行驶工况及汽车运动特征评估体系，分别计算出汽车行驶工况与该城市所采集数据源（经处理后的数据）的各指标（运动特征）值，
# 并说明你们所构建的汽车行驶工况的合理性。

import numpy as np
import pandas as pd
from data import get_data
import openpyxl

outwb = openpyxl.Workbook()  # 打开一个将写的文件
outws = outwb.create_sheet(index=0)  # 在将写的文件创建she

def read_data():
    save_data = []
    rubbish = []
    # save_v_1,a_1,a_0,dasu,a_1_100,a_0_100,a_std = [],[],[],[],[],[],[]
    save_v_1, a_1, a_0, a_std = [], [], [], []
    for n in range(713-1):
        x1= get_data(n,n+1)  # 提取所有列数据
        save_data.clear()
        rubbish.clear()
        save_v_1.clear()
        a_1.clear()
        a_0.clear()
        a_std.clear()
        a0 = 0
        for i in x1: # 提取每一列数据
            if (np.isnan(i)):
                rubbish.append(i)
            else:
                save_data.append(i[0])
        for k in save_data:
            if k>0:
                save_v_1.append(k)
            a_1.append(k-a0)
            a0 = k

        for j in a_1: # 加速度##############
            if j <0:
                a_0.append(j)

        dasu = 1 - round(len(save_v_1)/len(save_data),2) # 计算怠速####################
        a_0_100 = round(len(a_0)/len(save_data),2)  # 减速时间比
        a_1_100 = 1- a_0_100 # 加速时间比



        outws.cell(n+2, 1).value = np.mean(save_data)  # 写文件
        outws.cell(n+2, 2).value = np.mean(save_v_1)
        outws.cell(n + 2, 3).value = np.mean(a_1)  # 写文件
        outws.cell(n + 2, 4).value = np.mean(a_0)  # 写文件
        outws.cell(n + 2, 5).value = dasu  # 写文件
        outws.cell(n + 2, 6).value = a_1_100  # 写文件
        outws.cell(n + 2, 7).value = a_0_100  # 写文件
        outws.cell(n + 2, 8).value = np.std(save_data)  # 写文件
        outws.cell(n + 2, 9).value = np.std(a_1)  # 写文件
        print(np.mean(save_data))
    outws.cell(1, 1).value = "平均速度"  # 写文件
    outws.cell(1, 2).value = "平均行驶速度"  # 写文件
    outws.cell(1, 3).value = "平均加速度"  # 写文件
    outws.cell(1, 4).value = "平均减速度"  # 写文件
    outws.cell(1, 5).value = "怠速时间比"  # 写文件
    outws.cell(1, 6).value = "加速时间比"  # 写文件
    outws.cell(1, 7).value = "减速时间比"  # 写文件
    outws.cell(1, 8).value = "速度标准差"  # 写文件
    outws.cell(1, 9).value = "加速度标准差"  # 写文件


if __name__ == '__main__':
    read_data()
    outwb.save("./data_set/feature_1.xlsx")
    print('创建excel文件完成！')
